# sistemas/classificador_documentos/services_export.py
"""
Serviço de exportação de resultados de classificação.

Suporta exportação em:
- Excel (.xlsx)
- CSV

Autor: LAB/PGE-MS
"""

import io
import re
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime

from .models import ResultadoClassificacao, ExecucaoClassificacao

logger = logging.getLogger(__name__)


class ExportService:
    """Serviço para exportação de resultados de classificação"""

    # Caracteres ilegais em Excel
    ILLEGAL_CHARACTERS_RE = re.compile(
        r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]'
    )

    def _limpar_texto_excel(self, texto: str) -> str:
        """Remove caracteres ilegais para Excel"""
        if not texto:
            return ""
        return self.ILLEGAL_CHARACTERS_RE.sub('', str(texto))

    def exportar_excel(
        self,
        resultados: List[ResultadoClassificacao],
        execucao: Optional[ExecucaoClassificacao] = None
    ) -> io.BytesIO:
        """
        Exporta resultados para Excel.

        Args:
            resultados: Lista de resultados de classificação
            execucao: Execução associada (opcional, para metadados)

        Returns:
            BytesIO com o arquivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl não está instalado")

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Cabeçalhos
        headers = [
            "Código Documento",
            "Número Processo",
            "Nome Arquivo",
            "Categoria",
            "Subcategoria",
            "Confiança",
            "Justificativa",
            "Status",
            "Fonte",
            "Extração Via",
            "Tokens",
            "Erro",
            "Data Processamento"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Dados
        for row_idx, resultado in enumerate(resultados, 2):
            ws.cell(row=row_idx, column=1, value=self._limpar_texto_excel(resultado.codigo_documento))
            ws.cell(row=row_idx, column=2, value=self._limpar_texto_excel(resultado.numero_processo))
            ws.cell(row=row_idx, column=3, value=self._limpar_texto_excel(resultado.nome_arquivo))
            ws.cell(row=row_idx, column=4, value=self._limpar_texto_excel(resultado.categoria))
            ws.cell(row=row_idx, column=5, value=self._limpar_texto_excel(resultado.subcategoria))
            ws.cell(row=row_idx, column=6, value=self._limpar_texto_excel(resultado.confianca))
            ws.cell(row=row_idx, column=7, value=self._limpar_texto_excel(resultado.justificativa))
            ws.cell(row=row_idx, column=8, value=self._limpar_texto_excel(resultado.status))
            ws.cell(row=row_idx, column=9, value=self._limpar_texto_excel(resultado.fonte))
            ws.cell(row=row_idx, column=10, value=self._limpar_texto_excel(resultado.texto_extraido_via))
            ws.cell(row=row_idx, column=11, value=resultado.tokens_extraidos)
            ws.cell(row=row_idx, column=12, value=self._limpar_texto_excel(resultado.erro_mensagem))

            if resultado.processado_em:
                ws.cell(row=row_idx, column=13, value=resultado.processado_em.strftime("%d/%m/%Y %H:%M:%S"))

        # Ajusta largura das colunas
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)

            for row in range(1, min(len(resultados) + 2, 100)):  # Limita a 100 linhas para cálculo
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)  # Max 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width

        # Adiciona aba de metadados se tiver execução
        if execucao:
            ws_meta = wb.create_sheet("Metadados")
            ws_meta.cell(row=1, column=1, value="Propriedade")
            ws_meta.cell(row=1, column=2, value="Valor")

            metadados = [
                ("ID Execução", execucao.id),
                ("Status", execucao.status),
                ("Total Arquivos", execucao.total_arquivos),
                ("Processados", execucao.arquivos_processados),
                ("Sucesso", execucao.arquivos_sucesso),
                ("Erros", execucao.arquivos_erro),
                ("Modelo", execucao.modelo_usado),
                ("Iniciado em", execucao.iniciado_em.strftime("%d/%m/%Y %H:%M:%S") if execucao.iniciado_em else ""),
                ("Finalizado em", execucao.finalizado_em.strftime("%d/%m/%Y %H:%M:%S") if execucao.finalizado_em else ""),
                ("Exportado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            ]

            for row_idx, (prop, valor) in enumerate(metadados, 2):
                ws_meta.cell(row=row_idx, column=1, value=prop)
                ws_meta.cell(row=row_idx, column=2, value=str(valor) if valor else "")

        # Salva em BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def exportar_csv(
        self,
        resultados: List[ResultadoClassificacao],
        separador: str = ";"
    ) -> io.StringIO:
        """
        Exporta resultados para CSV.

        Args:
            resultados: Lista de resultados de classificação
            separador: Separador de campos (default: ;)

        Returns:
            StringIO com o arquivo CSV
        """
        import csv

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=separador, quoting=csv.QUOTE_ALL)

        # Cabeçalhos
        headers = [
            "codigo_documento",
            "numero_processo",
            "nome_arquivo",
            "categoria",
            "subcategoria",
            "confianca",
            "justificativa",
            "status",
            "fonte",
            "extracao_via",
            "tokens",
            "erro",
            "data_processamento"
        ]
        writer.writerow(headers)

        # Dados
        for resultado in resultados:
            row = [
                resultado.codigo_documento or "",
                resultado.numero_processo or "",
                resultado.nome_arquivo or "",
                resultado.categoria or "",
                resultado.subcategoria or "",
                resultado.confianca or "",
                resultado.justificativa or "",
                resultado.status or "",
                resultado.fonte or "",
                resultado.texto_extraido_via or "",
                resultado.tokens_extraidos or "",
                resultado.erro_mensagem or "",
                resultado.processado_em.strftime("%d/%m/%Y %H:%M:%S") if resultado.processado_em else ""
            ]
            writer.writerow(row)

        buffer.seek(0)
        return buffer

    def exportar_json(self, resultados: List[ResultadoClassificacao]) -> List[Dict[str, Any]]:
        """
        Exporta resultados para lista de dicionários (JSON).

        Args:
            resultados: Lista de resultados de classificação

        Returns:
            Lista de dicionários
        """
        return [
            {
                "codigo_documento": r.codigo_documento,
                "numero_processo": r.numero_processo,
                "nome_arquivo": r.nome_arquivo,
                "categoria": r.categoria,
                "subcategoria": r.subcategoria,
                "confianca": r.confianca,
                "justificativa": r.justificativa,
                "status": r.status,
                "fonte": r.fonte,
                "extracao_via": r.texto_extraido_via,
                "tokens": r.tokens_extraidos,
                "erro": r.erro_mensagem,
                "resultado_completo": r.resultado_json,
                "data_processamento": r.processado_em.isoformat() if r.processado_em else None
            }
            for r in resultados
        ]


# Instância global
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Retorna instância singleton do serviço de exportação"""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
